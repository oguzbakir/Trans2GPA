import pdftotext
import openpyxl

grades = {'A1': 12, 'A2': 11, 'A3': 10, 'B1': 9, 'B2': 8, 'B3': 7, 'C1': 6, 'C2': 5, 'C3': 4, 'D': 3, 'F1': 2, 'F2': 1,
          'F3': 0}

"""
:/
9 -> Technical Elective(non-lab)
10 -> Technical Elective(with-lab)
11 -> Technical Elective Lab.
:\
"""
lecture_semesters = {'BBM101': {'semester': 1, 'index': ['B', 20]}, 'BBM103': {'semester': 1, 'index': ['B', 21]},
                     'BBM105': {'semester': 1, 'index': ['B', 22]}, 'FİZ137': {'semester': 1, 'index': ['B', 23]},
                     'MAT123': {'semester': 1, 'index': ['B', 24]}, 'İNG111': {'semester': 1, 'index': ['B', 25]},
                     'TKD103': {'semester': 1, 'index': ['B', 26]}, 'BEB650': {'semester': 1, 'index': ['B', 27]},
                     'FİZ103': {'semester': 1, 'index': ['B', 28]}, 'ÜNİ101': {'semester': 1, 'index': ['B', 29]},
                     'BBM102': {'semester': 2, 'index': ['I', 20]}, 'BBM104': {'semester': 2, 'index': ['I', 21]},
                     'MAT124': {'semester': 2, 'index': ['I', 22]}, 'FİZ138': {'semester': 2, 'index': ['I', 23]},
                     'FİZ117': {'semester': 2, 'index': ['I', 24]}, 'İNG112': {'semester': 2, 'index': ['I', 25]},
                     'TKD104': {'semester': 2, 'index': ['I', 26]}, 'FİZ104': {'semester': 2, 'index': ['I', 27]},
                     'BBM201': {'semester': 3, 'index': ['B', 40]}, 'BBM203': {'semester': 3, 'index': ['B', 41]},
                     'BBM205': {'semester': 3, 'index': ['B', 42]}, 'BBM231': {'semester': 3, 'index': ['B', 43]},
                     'BBM233': {'semester': 3, 'index': ['B', 44]}, 'İST299': {'semester': 3, 'index': ['B', 45]},
                     'AİT203': {'semester': 3, 'index': ['B', 46]}, 'MÜH103': {'semester': 3, 'index': ['B', 47]},
                     'HAS222': {'semester': 3, 'index': ['B', 48]},
                     'BBM202': {'semester': 4, 'index': ['I', 40]}, 'BBM204': {'semester': 4, 'index': ['I', 41]},
                     'BBM234': {'semester': 4, 'index': ['I', 42]}, 'MAT254': {'semester': 4, 'index': ['I', 43]},
                     'ELE296': {'semester': 4, 'index': ['I', 44]}, 'İST292': {'semester': 4, 'index': ['I', 45]},
                     'AİT204': {'semester': 4, 'index': ['I', 46]}, 'MÜH104': {'semester': 4, 'index': ['I', 47]},
                     'HAS223': {'semester': 4, 'index': ['I', 48]},
                     'BBM301': {'semester': 5, 'index': ['B', 60]}, 'BBM325': {'semester': 5, 'index': ['B', 61]},
                     'BBM341': {'semester': 5, 'index': ['B', 62]}, 'BBM371': {'semester': 5, 'index': ['B', 63]},
                     'BBM342': {'semester': 6, 'index': ['I', 64]}, 'BBM382': {'semester': 6, 'index': ['I', 65]},
                     'BBM384': {'semester': 6, 'index': ['I', 66]},
                     'BBM425': {'semester': 7, 'index': ['B', 80]}, 'BBM427': {'semester': 7, 'index': ['B', 81]},
                     'BBM479': {'semester': 7, 'index': ['B', 82]}, 'BBM419': {'semester': 7, 'index': ['B', 83]},
                     'BBM428': {'semester': 8, 'index': ['I', 84]}, 'BBM480': {'semester': 8, 'index': ['I', 85]},
                     'BBM420': {'semester': 8, 'index': ['I', 86]},
                     'BBM401': {'semester': 9, 'index': []}, 'BBM402': {'semester': 9, 'index': []},
                     'BBM403': {'semester': 9, 'index': []}, 'BBM404': {'semester': 9, 'index': []},
                     'BBM405': {'semester': 9, 'index': []}, 'BBM407': {'semester': 9, 'index': []},
                     'BBM408': {'semester': 9, 'index': []},
                     'BBM410': {'semester': 9, 'index': []}, 'BBM411': {'semester': 9, 'index': []},
                     'BBM431': {'semester': 9, 'index': []}, 'BBM441': {'semester': 9, 'index': []},
                     'BBM442': {'semester': 9, 'index': []}, 'BBM443': {'semester': 9, 'index': []},
                     'BBM456': {'semester': 9, 'index': []},
                     'BBM461': {'semester': 9, 'index': []}, 'BBM462': {'semester': 9, 'index': []},
                     'BBM475': {'semester': 9, 'index': []}, 'BBM485': {'semester': 9, 'index': []},
                     'BBM486': {'semester': 9, 'index': []}, 'BBM498': {'semester': 9, 'index': []},
                     'BBM406': {'semester': 10, 'index': []}, 'BBM412': {'semester': 10, 'index': []},
                     'BBM413': {'semester': 10, 'index': []}, 'BBM416': {'semester': 10, 'index': []},
                     'BBM421': {'semester': 10, 'index': []}, 'BBM422': {'semester': 10, 'index': []},
                     'BBM432': {'semester': 10, 'index': []},
                     'BBM433': {'semester': 10, 'index': []}, 'BBM444': {'semester': 10, 'index': []},
                     'BBM451': {'semester': 10, 'index': []}, 'BBM452': {'semester': 10, 'index': []},
                     'BBM458': {'semester': 10, 'index': []}, 'BBM463': {'semester': 10, 'index': []},
                     'BBM467': {'semester': 10, 'index': []},
                     'BBM471': {'semester': 10, 'index': []}, 'BBM472': {'semester': 10, 'index': []},
                     'BBM481': {'semester': 10, 'index': []}, 'BBM482': {'semester': 10, 'index': []},
                     'BBM490': {'semester': 10, 'index': []}, 'BBM491': {'semester': 10, 'index': []},
                     'BBM492': {'semester': 10, 'index': []}, 'BBM495': {'semester': 10, 'index': []},
                     'BBM409': {'semester': 11, 'index': []}, 'BBM414': {'semester': 11, 'index': []},
                     'BBM415': {'semester': 11, 'index': []}, 'BBM418': {'semester': 11, 'index': []},
                     'BBM423': {'semester': 11, 'index': []}, 'BBM424': {'semester': 11, 'index': []},
                     'BBM434': {'semester': 11, 'index': []},
                     'BBM436': {'semester': 11, 'index': []}, 'BBM446': {'semester': 11, 'index': []},
                     'BBM453': {'semester': 11, 'index': []}, 'BBM459': {'semester': 11, 'index': []},
                     'BBM460': {'semester': 11, 'index': []}, 'BBM465': {'semester': 11, 'index': []},
                     'BBM469': {'semester': 11, 'index': []},
                     'BBM474': {'semester': 11, 'index': []}, 'BBM483': {'semester': 11, 'index': []},
                     'BBM484': {'semester': 11, 'index': []}, 'BBM488': {'semester': 11, 'index': []},
                     'BBM493': {'semester': 11, 'index': []}, 'BBM494': {'semester': 11, 'index': []},
                     'BBM497': {'semester': 11, 'index': []}}
"""
Every index shows each semester, in each index:
increase first index if there is a non-technical elective
increase second index if there is a technical elective
increase third index if there is a technical elective with lab
"""
electives = [[0, 0, 0], [0, 0, 0], [1, 0, 0], [1, 0, 0], [1, 0, 1], [1, 2, 0], [1, 1, 1], [1, 0, 2]]


def get_lectures(pdf_name):
    with open(pdf_name, "rb") as f:
        pdf = pdftotext.PDF(f)
        lectures = {}
        for page in pdf:
            page_text = page.split('\n')

            for line in page_text:
                clean_line = line.strip()
                if clean_line != '' and clean_line[0] == '│':  # Get useful lines from table
                    clean_line_list = list(filter(None, clean_line.split(' ')))
                    if len(clean_line_list) > 1:  # Eliminate garbage lines
                        grade = clean_line_list[-1].replace('│', '')
                        if grade in grades:  # Check if line is a lecture line
                            code = clean_line_list[1]
                            if code not in lectures:  # Fresh lecture
                                lectures[code] = {'grade': grade,
                                                  'name': ' '.join(clean_line_list[2:-4]),
                                                  'akts': int(clean_line_list[-3])}
                            else:  # Repeated lecture
                                if grades[lectures[code]['grade']] < grades[grade]:  # New grade is bigger, change it
                                    lectures[code] = {'grade': grade,
                                                      'name': ' '.join(clean_line_list[2:-4]),
                                                      'akts': int(clean_line_list[-3])}

    return lectures


lectures = get_lectures('start.pdf')
wb = openpyxl.load_workbook(filename='gpa.xlsx')
ws = wb.active
for key in lectures.keys():
    try:
        lecture = lecture_semesters[key]
        if lecture['semester'] not in [9, 10, 11]:
            ws[chr(ord(lecture['index'][0])) + str(lecture['index'][1])] = key
            ws[chr(ord(lecture['index'][0]) + 1) + str(lecture['index'][1])] = lectures[key]['name']
            ws[chr(ord(lecture['index'][0]) + 2) + str(lecture['index'][1])] = lectures[key]['akts']
            ws[chr(ord(lecture['index'][0]) + 3) + str(lecture['index'][1])] = lectures[key]['grade']
        elif lecture['semester'] == 9:
            pass
        elif lecture['semester'] == 10:
            pass
        elif lecture['semester'] == 11:
            pass
    except KeyError:
        print("{0} coded lecture not added".format(key))

wb.save('out.xlsx')